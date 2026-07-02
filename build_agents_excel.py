"""One-shot script: config/agents.json → docs/agents_overview.xlsx"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

ROOT = Path(__file__).parent
data = json.loads((ROOT / "config" / "agents.json").read_text(encoding="utf-8"))
agents = data["agents"]

wb = openpyxl.Workbook()

# ── helpers ────────────────────────────────────────────────────────────────

def hdr_style(cell, bg="2E4057"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

LAYER_BG = {1: "FFE0E0", 2: "FFE8CC", 3: "FFFACC", 4: "DDEEDD", 5: "DDE8FF"}
LAYER_HDR = {1: "C0392B", 2: "D35400", 3: "B7950B", 4: "1E8449", 5: "1A5276"}

EXPERTISE = [
    "consulting", "training", "coaching", "engineering", "operations",
    "marketing", "finance", "design", "sales", "leadership",
    "technology", "healthcare", "education", "legal", "real_estate",
]

# ── Sheet 1: Agent Overview ─────────────────────────────────────────────────

ws = wb.active
ws.title = "Agent Overview"

COLS = [
    ("Layer",           8),
    ("Action Type",    26),
    ("Label",          46),
    ("Description",    62),
    ("Model",          10),
    ("Cold Start?",    12),
    ("CS Order",        9),
    ("CS Prior",        9),
    ("Hub Node?",      11),
    ("Hub Label",      34),
    ("Disclaimer",     11),
    ("Large Output",   13),
    ("Integrations",   34),
    ("Prerequisites",  34),
    ("Triggered By",   30),
    ("Data Dependencies", 50),
    ("Prompt Fields",  55),
    ("Positive Signals", 38),
    ("Negative Signals", 28),
]

for col_idx, (name, width) in enumerate(COLS, 1):
    cell = ws.cell(row=1, column=col_idx, value=name)
    hdr_style(cell)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 32

for row_idx, agent in enumerate(agents, 2):
    layer = agent.get("layer", 0)
    bg = LAYER_BG.get(layer, "FFFFFF")
    fill = PatternFill("solid", fgColor=bg)
    border = thin_border()

    prereqs = ", ".join(p["action_type"] for p in agent.get("prerequisites", []))
    prompt_fields = ", ".join(
        f"{p['key']} ({p['type']})" for p in agent.get("prompt_form", [])
    )

    row_data = [
        layer,
        agent.get("action_type", ""),
        agent.get("label", ""),
        agent.get("description", ""),
        agent.get("model", ""),
        "Yes" if agent.get("cold_start_chain") else "No",
        agent.get("cold_start_order", ""),
        agent.get("cold_start_prior", ""),
        "Yes" if agent.get("hub_node") else "No",
        agent.get("hub_label", ""),
        "Yes" if agent.get("disclaimer") else "No",
        "Yes" if agent.get("large_output") else "No",
        ", ".join(agent.get("integrations", [])),
        prereqs,
        ", ".join(agent.get("triggered_by", [])),
        ", ".join(agent.get("data_dependencies", [])),
        prompt_fields,
        ", ".join(agent.get("signals", {}).get("positive", [])),
        ", ".join(agent.get("signals", {}).get("negative", [])),
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[row_idx].height = 60

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Sheet 2: Expertise Relevance ────────────────────────────────────────────

ws2 = wb.create_sheet("Expertise Relevance")

er_cols = ["Layer", "Action Type", "Label"] + [d.title().replace("_", " ") for d in EXPERTISE]
for col_idx, name in enumerate(er_cols, 1):
    cell = ws2.cell(row=1, column=col_idx, value=name)
    hdr_style(cell)

ws2.column_dimensions["A"].width = 8
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 46
for col_idx in range(4, len(er_cols) + 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 13

ws2.row_dimensions[1].height = 32

for row_idx, agent in enumerate(agents, 2):
    er = agent.get("expertise_relevance", {})
    row_data = [
        agent.get("layer", ""),
        agent.get("action_type", ""),
        agent.get("label", ""),
    ] + [er.get(d, "") for d in EXPERTISE]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border()

        if col_idx > 3 and isinstance(value, (int, float)):
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            if value >= 0.85:
                cell.fill = PatternFill("solid", fgColor="00B050")
                cell.font = Font(bold=True, color="FFFFFF")
            elif value >= 0.70:
                cell.fill = PatternFill("solid", fgColor="92D050")
            elif value >= 0.55:
                cell.fill = PatternFill("solid", fgColor="FFFF00")
            else:
                cell.fill = PatternFill("solid", fgColor="FF9999")
        else:
            cell.alignment = Alignment(vertical="top")

ws2.freeze_panes = "D2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(er_cols))}1"

# ── Sheet 3: Layer Summary ──────────────────────────────────────────────────

ws3 = wb.create_sheet("Layer Summary")

LAYER_META = {
    1: "Consulting & Client Acquisition",
    2: "Group Delivery & Training",
    3: "Digital Products",
    4: "Content & Audience",
    5: "Wealth & Asset Protection",
}

s3_headers = ["Layer", "Theme", "# Agents", "Cold-Start Agents", "Hub Nodes",
              "Agents with Disclaimer", "Avg CS Prior", "Top Integration"]
for col_idx, name in enumerate(s3_headers, 1):
    cell = ws3.cell(row=1, column=col_idx, value=name)
    hdr_style(cell)
    ws3.column_dimensions[get_column_letter(col_idx)].width = [
        8, 38, 10, 18, 12, 22, 14, 22
    ][col_idx - 1]

ws3.row_dimensions[1].height = 32

for row_idx, layer_num in enumerate(range(1, 6), 2):
    layer_agents = [a for a in agents if a.get("layer") == layer_num]
    cs_agents = [a for a in layer_agents if a.get("cold_start_chain")]
    hub_agents = [a for a in layer_agents if a.get("hub_node")]
    disclaimer_agents = [a for a in layer_agents if a.get("disclaimer")]
    cs_priors = [a.get("cold_start_prior", 0) for a in layer_agents if a.get("cold_start_prior")]
    avg_prior = round(sum(cs_priors) / len(cs_priors), 2) if cs_priors else 0

    all_integrations = []
    for a in layer_agents:
        all_integrations.extend(a.get("integrations", []))
    top_integration = max(set(all_integrations), key=all_integrations.count) if all_integrations else "—"

    bg = LAYER_BG.get(layer_num, "FFFFFF")
    fill = PatternFill("solid", fgColor=bg)

    row_data = [
        layer_num,
        LAYER_META.get(layer_num, ""),
        len(layer_agents),
        len(cs_agents),
        len(hub_agents),
        len(disclaimer_agents),
        avg_prior,
        top_integration,
    ]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center" if col_idx != 2 else "left", vertical="center")
        if col_idx == 2:
            cell.font = Font(bold=True)

# ── Sheet 4: Prompt Forms ───────────────────────────────────────────────────

ws4 = wb.create_sheet("Prompt Forms")

pf_headers = ["Layer", "Action Type", "Label", "Field Key", "Field Label", "Type", "Required", "Options"]
for col_idx, name in enumerate(pf_headers, 1):
    cell = ws4.cell(row=1, column=col_idx, value=name)
    hdr_style(cell)

ws4.column_dimensions["A"].width = 8
ws4.column_dimensions["B"].width = 26
ws4.column_dimensions["C"].width = 42
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 40
ws4.column_dimensions["F"].width = 12
ws4.column_dimensions["G"].width = 10
ws4.column_dimensions["H"].width = 50
ws4.row_dimensions[1].height = 32

pf_row = 2
for agent in agents:
    layer = agent.get("layer", 0)
    bg = LAYER_BG.get(layer, "FFFFFF")
    fill = PatternFill("solid", fgColor=bg)
    for field in agent.get("prompt_form", []):
        row_data = [
            layer,
            agent.get("action_type", ""),
            agent.get("label", ""),
            field.get("key", ""),
            field.get("label", ""),
            field.get("type", ""),
            "Yes" if field.get("required") else "No",
            ", ".join(field.get("options", [])),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=pf_row, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        pf_row += 1

ws4.freeze_panes = "A2"
ws4.auto_filter.ref = f"A1:{get_column_letter(len(pf_headers))}1"

# ── Save ────────────────────────────────────────────────────────────────────

out = ROOT / "docs" / "agents_overview.xlsx"
wb.save(out)
print(f"Saved: {out}")
